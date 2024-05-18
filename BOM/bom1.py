from openpyxl import Workbook,load_workbook

#这个是判断（导出bom）ws1中的某列和（系统物料表）ws2中的某列是否匹配，如果匹配，则把ws2的行内容赋值到ws3,不匹配把ws1的行内容赋值到ws3

wb1 = load_workbook(r"D:\Code_python\exercise\openpyxl\test.xlsx")
wb2 = load_workbook(r"D:\Code_python\exercise\openpyxl\test1.xlsx")
wb3 = load_workbook(r"D:\Code_python\exercise\openpyxl\test2.xlsx")
#设置sheet
ws1 = wb1["1she"]
ws2 = wb2["a"]
ws3 = wb3["Sheet1"]


#这个是行单元格复制函数:把src_sheet的src_cell那行开始往后的单元复制给到dst_sheet的dst_cell(从第src_cell开始复制到dst_cell)
def copy_part_row(src_sheet,src_cell,dst_sheet,dst_cell):
    #找出需要复制的src_sheet的单元格范围
    for row in src_sheet.iter_rows(min_row=src_cell.row, max_row=src_cell.row, min_col=src_cell.column, max_col=src_sheet.max_column):
        for src_cell in row:
            dst_cell.value = src_cell.value              #赋值
            dst_cell = dst_cell.offset(row=0, column=1)  #右移一位

#这个是行单元格复制函数:把src_sheet的src_cell那一整行复制给到dst_sheet的dst_cell那一行(从第一个开始复制整行)
def copy_total_row(src_sheet,src_cell,dst_sheet,dst_cell):
    #找到对应的第一个单元格
    src_row = src_cell.row
    dst_row = dst_cell.row
    src_cell = src_sheet["A%d" %(src_row)]     
    dst_cell = dst_sheet["A%d" %(dst_row)]   
    copy_part_row(src_sheet,src_cell,dst_sheet,dst_cell)

    """
    for row in src_sheet.iter_rows(min_row=src_cell.row, max_row=src_cell.row, min_col=src_cell.column, max_col=src_sheet.max_column):
        for src_cell in row:
            dst_cell.value = src_cell.value              #赋值
            dst_cell = dst_cell.offset(row=0, column=1)  #右移一位
    """

#copy_row(ws1,ws1["A3"],ws3,ws3["B4"])  #copy函数测试


#定义数组
list_coordinate = []
list_value = []
#找出需要源文件需要根据哪列来进行匹配，赋值给到数组
col_ws1 = int(input("输入(数字)在src_sheet需要匹配的那一列:"))
for col in ws1.iter_cols(min_row=1,max_row=ws1.max_row,min_col=col_ws1,max_col=col_ws1):
    for cell in col:
        list_coordinate.append(cell.coordinate)
        list_value.append (cell.value)

print("list_coordinate[] = " + str(list_coordinate))
print("list_value[] = " + str(list_value))


#判断和ws2中col_ws2列是否找到匹配
col_ws2 = int(input("输入（数字）需要在系统文件匹配的那一列:"))    
i = 0
match = [0 for i in range(ws1.max_row)]  #定义标记数组，匹配标记为1
#判断是否匹配
for col in ws2.iter_cols(min_row=1,max_row=ws2.max_row,min_col=col_ws2,max_col=col_ws2):    
    for i in range(0,ws1.max_row):
        for cell in col: 
        # print(cell.value)
        # print(list_value[i])
            if list_value[i] == cell.value:
                print(str(list[i])+"match")
                #print(cell.value)
                # copy_part_row(ws2,cell,ws3,ws3[list_coordinate[i]])
                copy_total_row(ws2, cell, ws3, ws3[list_coordinate[i]])
                match[i] = 1
                break    #找到第一个匹配后退出，一般只会有唯一匹配
        
        if match[i] == 0 :
            print(str(list[i])+"mismatch")
            #print(cell.value)
            # copy_part_row(ws1,ws1[list_coordinate[i]],ws3,ws3[list_coordinate[i]])
            copy_total_row(ws1, ws1[list_coordinate[i]], ws3, ws3[list_coordinate[i]])
        i = i+1

print("match[] = " +str(match))

wb3.save(r"D:\Code_python\exercise\openpyxl\test2.xlsx")