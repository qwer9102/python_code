from openpyxl import Workbook,load_workbook

#这个是判断（导出bom）ws1中的某列和（系统物料表）ws2中的某列是否匹配，如果匹配，则把ws2的行内容赋值到ws3,不匹配把ws1的行内容赋值到ws3

wb1 = load_workbook(r"D:\Code_python\exercise\openpyxl\word1.xlsx")

#设置sheet
ws1 = wb1["Sheet1"]
ws2 = wb1["Sheet2"]
ws3 = wb1["Sheet3"]


"""
# 1.字符串分割,先把需要分割的值赋值给到list_value/coordinate_spi数组
list_coordinate_split = []
list_value_split = []

col_ws2 = int(input("输入(数字)在search_sheet需要分割的那一列:"))
for col in ws2.iter_cols(min_row=1,max_row=ws2.max_row,min_col=col_ws2,max_col=col_ws2):
    for cell in col:
        #list_coordinate_split.append(cell.coordinate)
        list_value_split.append (cell.value)

# print("list_coordinate_split[] = " + str(list_coordinate_split))
# print("list_value_split[] = " + str(list_value_split))
# print(list_value_split[1])
# print("\n\n")

# 1.1开始分割
list_split = []
i = 1
print(list_value_split[i])
print("\n\n")
for i in range(1,ws2.max_row+1):
    str = list_value_split[i-1]
    if(str == None): 
        i = i+1                         #没有分割点的直接整个保存
    else:
        print(str)
        # list_split = str.split('[')   #按照[进行分割
        list_split = str.split('(')     #按照（进行分割
        # print(list_split)
        # print(list_split[0])          
        # print(list_split[1])
        ws2["A%d" %(i)] = list_split[0]     #分割后保存的位置，这里是保存到对应的第一列
        # ws2["B%d" %(i)] = list_split[1]   #这个原本是把分割后半部分分开保存的,但是有时是空会报错

        # print(ws2["A%d" %(i)].value)
        # print(ws2["B%d" %(i)].value)
        i = i+1

print("\n\n") 
print("split done!!! \n")     #正常分割完成标志
"""


""" """
#2.1.这个是部分行单元格复制函数:把src_sheet的src_cell那行开始往后的单元复制给到dst_sheet的dst_cell(从第src_cell开始复制到dst_cell)
def copy_part_row(src_sheet,src_cell,dst_sheet,dst_cell):
    #找出需要复制的src_sheet的单元格范围
    for row in src_sheet.iter_rows(min_row=src_cell.row, max_row=src_cell.row, min_col=src_cell.column, max_col=src_sheet.max_column):
        for src_cell in row:
            dst_cell.value = src_cell.value              #赋值
            dst_cell = dst_cell.offset(row=0, column=1)  #右移一位

#2.2.这个是整行单元格复制函数:把src_sheet的src_cell那一整行复制给到dst_sheet的dst_cell那一行(从第一个开始复制整行)
def copy_total_row(src_sheet,src_cell,dst_sheet,dst_cell):
    #找到对应的第一个单元格
    src_row = src_cell.row
    dst_row = dst_cell.row
    src_cell = src_sheet["A%d" %(src_row)]     
    dst_cell = dst_sheet["A%d" %(dst_row)]   
    copy_part_row(src_sheet,src_cell,dst_sheet,dst_cell)

"""
#2.3赋值的另一种实现方式
    for row in src_sheet.iter_rows(min_row=src_cell.row, max_row=src_cell.row, min_col=src_cell.column, max_col=src_sheet.max_column):
        for src_cell in row:
            dst_cell.value = src_cell.value              #赋值
            dst_cell = dst_cell.offset(row=0, column=1)  #右移一位
"""

#copy_row(ws1,ws1["A3"],ws3,ws3["B4"])  #copy函数测试


""""""
#3.复制前操作：先把需要src_sheet中需要匹配的赋值的字符串赋值给到list_coordinate/list_value数组
#声明数组
list_coordinate = []
list_value = []
#3.1.找出需要源文件需要根据哪列来进行匹配，赋值给到数组list_coordinate/list_value
col_ws1 = int(input("输入(数字)在src_sheet需要匹配的那一列:"))
for col in ws1.iter_cols(min_row=1,max_row=ws1.max_row,min_col=col_ws1,max_col=col_ws1):
    for cell in col:
        list_coordinate.append(cell.coordinate)
        list_value.append (cell.value)
#3.1.1赋值打印确认
# print("list_coordinate[] = " + str(list_coordinate))
print("list_value[] = " + str(list_value))
print("\n")


#3.2.和系统文件匹配
#判断和ws2中col_ws2列是否找到匹配
col_ws2 = int(input("输入（数字）需要在系统文件匹配的那一列:"))
    
i = 0
match = [0 for i in range(ws1.max_row)]    #定义标记数组并初始化为0，如果后面能匹配则修改为1
#3.2.1判断是否匹配
for col in ws2.iter_cols(min_row=1,max_row=ws2.max_row,min_col=col_ws2,max_col=col_ws2):    
    for i in range(0,ws1.max_row):
        for cell in col: 
        # print(cell.value)
        # print(list_value[i])
            if list_value[i] == cell.value:
                # print(str(list[i])+"match")
                # print(cell.value)
                # copy_part_row(ws2,cell,ws3,ws3[list_coordinate[i]])
                copy_total_row(ws2, cell, ws3, ws3[list_coordinate[i]])
                match[i] = 1
                break    #找到第一个匹配后退出，一般只会有唯一匹配
        
        if match[i] == 0 :
            # print(str(list[i])+"mismatch")
            #print(cell.value)
            # copy_part_row(ws1,ws1[list_coordinate[i]],ws3,ws3[list_coordinate[i]])
            copy_total_row(ws1, ws1[list_coordinate[i]], ws3, ws3[list_coordinate[i]])
        i = i+1

print("match[] = " +str(match))

# wb1.save(r"D:\Code_python\exercise\openpyxl\word1.xlsx")
